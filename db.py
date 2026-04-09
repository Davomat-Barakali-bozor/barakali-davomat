import os
import sqlite3
from datetime import datetime, date
from io import BytesIO
from zoneinfo import ZoneInfo
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font

DB_PATH = os.getenv('DB_PATH', '/data/davomat.db')
TIMEZONE = os.getenv('TIMEZONE', 'Asia/Tashkent')
WORK_START = os.getenv('WORK_START', '09:00')


def now_local() -> datetime:
    return datetime.now(ZoneInfo(TIMEZONE))


def get_conn() -> sqlite3.Connection:
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            user_id INTEGER PRIMARY KEY,
            full_name TEXT NOT NULL,
            username TEXT,
            phone TEXT,
            status TEXT DEFAULT 'pending',
            created_at TEXT,
            updated_at TEXT
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS employee_state (
            user_id INTEGER PRIMARY KEY,
            step TEXT,
            action TEXT,
            full_name TEXT,
            phone TEXT,
            photo_file_id TEXT,
            updated_at TEXT
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS admin_state (
            admin_id INTEGER PRIMARY KEY,
            step TEXT,
            updated_at TEXT
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            full_name TEXT NOT NULL,
            work_date TEXT NOT NULL,
            action TEXT NOT NULL,
            event_time TEXT NOT NULL,
            latitude REAL,
            longitude REAL,
            photo_file_id TEXT,
            created_at TEXT
        )
    ''')
    conn.commit()
    conn.close()


# ---------------- EMPLOYEE STATE ----------------
def set_employee_state(user_id: int, step: Optional[str], action: Optional[str] = None,
                       full_name: Optional[str] = None, phone: Optional[str] = None,
                       photo_file_id: Optional[str] = None) -> None:
    conn = get_conn()
    cur = conn.cursor()
    now = now_local().isoformat()

    if step is None:
        cur.execute('DELETE FROM employee_state WHERE user_id=?', (user_id,))
    else:
        cur.execute('SELECT * FROM employee_state WHERE user_id=?', (user_id,))
        row = cur.fetchone()
        if row:
            cur.execute(
                '''
                UPDATE employee_state
                SET step=?,
                    action=COALESCE(?, action),
                    full_name=COALESCE(?, full_name),
                    phone=COALESCE(?, phone),
                    photo_file_id=COALESCE(?, photo_file_id),
                    updated_at=?
                WHERE user_id=?
                ''',
                (step, action, full_name, phone, photo_file_id, now, user_id),
            )
        else:
            cur.execute(
                '''
                INSERT INTO employee_state (user_id, step, action, full_name, phone, photo_file_id, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ''',
                (user_id, step, action, full_name, phone, photo_file_id, now),
            )
    conn.commit()
    conn.close()


def get_employee_state(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT * FROM employee_state WHERE user_id=?', (user_id,))
    row = cur.fetchone()
    conn.close()
    return row


# ---------------- ADMIN STATE ----------------
def set_admin_state(admin_id: int, step: Optional[str]) -> None:
    conn = get_conn()
    cur = conn.cursor()
    now = now_local().isoformat()
    if step is None:
        cur.execute('DELETE FROM admin_state WHERE admin_id=?', (admin_id,))
    else:
        cur.execute('SELECT * FROM admin_state WHERE admin_id=?', (admin_id,))
        row = cur.fetchone()
        if row:
            cur.execute('UPDATE admin_state SET step=?, updated_at=? WHERE admin_id=?', (step, now, admin_id))
        else:
            cur.execute('INSERT INTO admin_state (admin_id, step, updated_at) VALUES (?, ?, ?)', (admin_id, step, now))
    conn.commit()
    conn.close()


def get_admin_state(admin_id: int) -> Optional[str]:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT step FROM admin_state WHERE admin_id=?', (admin_id,))
    row = cur.fetchone()
    conn.close()
    return row['step'] if row else None


# ---------------- EMPLOYEES ----------------
def upsert_pending_employee(user_id: int, full_name: str, username: Optional[str], phone: str) -> None:
    conn = get_conn()
    cur = conn.cursor()
    now = now_local().isoformat()
    cur.execute('SELECT * FROM employees WHERE user_id=?', (user_id,))
    row = cur.fetchone()
    if row:
        cur.execute(
            '''
            UPDATE employees
            SET full_name=?, username=?, phone=?, status='pending', updated_at=?
            WHERE user_id=?
            ''',
            (full_name, username, phone, now, user_id),
        )
    else:
        cur.execute(
            '''
            INSERT INTO employees (user_id, full_name, username, phone, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, 'pending', ?, ?)
            ''',
            (user_id, full_name, username, phone, now, now),
        )
    conn.commit()
    conn.close()


def approve_employee(user_id: int) -> bool:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('UPDATE employees SET status=?, updated_at=? WHERE user_id=?', ('approved', now_local().isoformat(), user_id))
    ok = cur.rowcount > 0
    conn.commit()
    conn.close()
    return ok


def remove_employee(user_id: int) -> bool:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('UPDATE employees SET status=?, updated_at=? WHERE user_id=?', ('removed', now_local().isoformat(), user_id))
    ok = cur.rowcount > 0
    conn.commit()
    conn.close()
    return ok


def get_employee(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT * FROM employees WHERE user_id=?', (user_id,))
    row = cur.fetchone()
    conn.close()
    return row


def is_employee_approved(user_id: int) -> bool:
    row = get_employee(user_id)
    return bool(row and row['status'] == 'approved')


def get_pending_employees():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT user_id, full_name, username, phone FROM employees WHERE status=? ORDER BY created_at DESC', ('pending',))
    rows = cur.fetchall()
    conn.close()
    return rows


def get_approved_employees():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT user_id, full_name, username, phone FROM employees WHERE status=? ORDER BY full_name ASC', ('approved',))
    rows = cur.fetchall()
    conn.close()
    return rows


# ---------------- ATTENDANCE ----------------
def record_attendance(user_id: int, action: str, latitude: float, longitude: float, photo_file_id: str):
    employee = get_employee(user_id)
    full_name = employee['full_name'] if employee else 'Noma’lum'
    now = now_local()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        '''
        INSERT INTO attendance (user_id, full_name, work_date, action, event_time, latitude, longitude, photo_file_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        (user_id, full_name, str(now.date()), action, now.isoformat(), latitude, longitude, photo_file_id, now.isoformat()),
    )
    conn.commit()
    conn.close()
    return full_name, now


def get_latest_today_action(user_id: int) -> Optional[str]:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        '''
        SELECT action FROM attendance
        WHERE user_id=? AND work_date=?
        ORDER BY id DESC LIMIT 1
        ''',
        (user_id, str(now_local().date())),
    )
    row = cur.fetchone()
    conn.close()
    return row['action'] if row else None


# ---------------- REPORTS ----------------
def create_daily_excel(work_date: date) -> BytesIO:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        '''
        SELECT e.user_id, e.full_name, e.username, e.phone,
               MIN(CASE WHEN a.action='checkin' THEN a.event_time END) AS first_in,
               MAX(CASE WHEN a.action='checkout' THEN a.event_time END) AS last_out
        FROM employees e
        LEFT JOIN attendance a ON a.user_id=e.user_id AND a.work_date=?
        WHERE e.status='approved'
        GROUP BY e.user_id, e.full_name, e.username, e.phone
        ORDER BY e.full_name ASC
        ''',
        (str(work_date),),
    )
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kunlik hisobot'
    headers = ['Sana', 'Xodim', 'Username', 'Telefon', 'Kelgan', 'Ketgan', 'Holat']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        first_in = row['first_in']
        last_out = row['last_out']
        if first_in and last_out:
            status = 'Kelgan/Ketgan'
        elif first_in:
            status = 'Kelgan'
        else:
            status = 'Kelmagan'
        ws.append([
            str(work_date),
            row['full_name'],
            f"@{row['username']}" if row['username'] else '-',
            row['phone'] or '-',
            first_in[11:16] if first_in else '-',
            last_out[11:16] if last_out else '-',
            status,
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def create_month_excel(year: int, month: int) -> BytesIO:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        '''
        SELECT e.full_name,
               COUNT(DISTINCT CASE WHEN a.action='checkin' THEN a.work_date END) AS days_in,
               COUNT(DISTINCT CASE WHEN a.action='checkout' THEN a.work_date END) AS days_out
        FROM employees e
        LEFT JOIN attendance a ON a.user_id=e.user_id AND substr(a.work_date, 1, 7)=?
        WHERE e.status='approved'
        GROUP BY e.full_name
        ORDER BY e.full_name ASC
        ''',
        (f'{year:04d}-{month:02d}',),
    )
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Oylik hisobot'
    headers = ['Xodim', 'Kelgan kun', 'Ketgan kun']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row['full_name'], row['days_in'], row['days_out']])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
